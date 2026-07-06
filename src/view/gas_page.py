from __future__ import annotations

import math
import re
from datetime import datetime
from pathlib import Path
from time import monotonic

import pyqtgraph as pg
from PyQt6.QtCore import QCoreApplication, QDate, QRect, Qt
from PyQt6.QtGui import QAction, QColor, QFont, QPainter, QPen, QPixmap
from PyQt6.QtWidgets import (
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from src.common.app_paths import app_root
from src.store.gas_store import GasStore
from src.view.components import (
    AntdDateRangePicker,
    GripSplitterHandle,
    NumericKeypad,
    VerticalProgressBar,
)


_orig_translate = QCoreApplication.translate


def _chinese_translate(context, src, disambiguation=None, n=-1):
    trans_map = {
        "View All": "显示全部",
        "X axis": "X 轴",
        "Y axis": "Y 轴",
        "Mouse Mode": "鼠标模式",
        "Plot Options": "绘图选项",
        "Export...": "导出数据...",
        "Set Safety Threshold...": "设置安全阈值...",
    }
    return trans_map.get(src, _orig_translate(context, src, disambiguation, n))


QCoreApplication.translate = _chinese_translate


class GroupHeaderWidget(QWidget):
    def __init__(self, table, groups, parent=None):
        super().__init__(parent)
        self.table = table
        self.groups = groups
        self.setFixedHeight(56)
        self.setMouseTracking(True)
        self.setStyleSheet("background: white;")

        self.table.horizontalHeader().sectionResized.connect(self.update)
        self.table.horizontalScrollBar().valueChanged.connect(self.update)

    def _offset_x(self):
        viewport = self.table.viewport()
        return viewport.mapToGlobal(viewport.rect().topLeft()).x() - self.mapToGlobal(
            self.rect().topLeft()
        ).x()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        header = self.table.horizontalHeader()
        offset_x = self._offset_x()

        painter.setPen(QPen(QColor("#F0F0F0"), 1))
        painter.setBrush(QColor("white"))
        painter.drawRect(self.rect())

        font_main = QFont("Microsoft YaHei", 9)
        font_main.setBold(True)
        font_sub = QFont("Microsoft YaHei", 8)

        current_col = 0
        for group in self.groups:
            label = group["label"]
            span = group["span"]
            subs = group.get("subs", [])

            x_pos = header.sectionViewportPosition(current_col) + offset_x
            width = sum(header.sectionSize(current_col + i) for i in range(span))
            rect = QRect(x_pos, 0, width, self.height())

            if rect.right() > 0 and rect.left() < self.width():
                if not subs:
                    painter.setFont(font_main)
                    painter.setPen(QColor("#333"))
                    painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, label)
                else:
                    top_rect = QRect(x_pos, 0, width, self.height() // 2)
                    painter.setFont(font_main)
                    painter.setPen(QColor("#666"))
                    painter.drawText(top_rect, Qt.AlignmentFlag.AlignCenter, label)

                    sub_x = x_pos
                    painter.setFont(font_sub)
                    painter.setPen(QColor("#999"))
                    for index, sub_label in enumerate(subs):
                        sub_w = header.sectionSize(current_col + index)
                        sub_rect = QRect(sub_x, self.height() // 2, sub_w, self.height() // 2)
                        painter.drawText(sub_rect, Qt.AlignmentFlag.AlignCenter, sub_label)
                        if index < len(subs) - 1:
                            painter.setPen(QColor("#F5F5F5"))
                            painter.drawLine(
                                sub_x + sub_w,
                                self.height() // 2 + 6,
                                sub_x + sub_w,
                                self.height() - 6,
                            )
                            painter.setPen(QColor("#999"))
                        sub_x += sub_w

                    painter.setPen(QColor("#F0F0F0"))
                    painter.drawLine(
                        x_pos + 4, self.height() // 2, x_pos + width - 4, self.height() // 2
                    )

                painter.setPen(QPen(QColor("#F0F0F0"), 1))
                if current_col + span < self.table.columnCount():
                    painter.drawLine(rect.right(), 10, rect.right(), self.height() - 10)

            current_col += span

        painter.setPen(QPen(QColor("#E8E8E8"), 1))
        painter.drawLine(0, self.height() - 1, self.width(), self.height() - 1)


class GasPage(QWidget):
    PRIMARY = "#597EF7"
    OK_GREEN = "#389E0D"
    NG_RED = "#D4380D"
    PROGRESS_NG = "#CF1322"
    GREY_BG = "#C4C4C4"
    IMAGE_COLUMN = 21

    def __init__(
        self,
        store: GasStore | None = None,
        manual_store=None,
        project_root: str | Path | None = None,
        parent=None,
    ):
        super().__init__(parent)
        self.store = store
        self.manual_store = manual_store
        self._project_root = Path(project_root) if project_root is not None else app_root()
        self.threshold = 85.0
        self.start_date = QDate.currentDate().addDays(-7)
        self.end_date = QDate.currentDate()
        self._last_query_state = None
        self._last_count_state = None
        # 工序曲线采集状态机: IDLE / COLLECTING / FROZEN
        self._chart_state = "IDLE"
        self._chart_max_samples = 500
        self._chart_started_at: float | None = None
        self._chart_x: list[float] = []
        self._chart_torque_1: list[float] = []
        self._chart_torque_2: list[float] = []
        self.setStyleSheet("background: transparent;")
        self._build_ui()
        self._connect_signals()

        if self.store and not any(self.store.date_range):
            self._sync_search_to_store()
        else:
            self._update_ui_from_store()

    def _build_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setHandleWidth(12)
        splitter.setStyleSheet("QSplitter::handle { background: transparent; }")
        splitter.addWidget(self._build_table_panel())
        splitter.addWidget(self._build_chart_panel())
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        handle = splitter.handle(1)
        grip = GripSplitterHandle(Qt.Orientation.Vertical, handle)
        grip_layout = QVBoxLayout(handle)
        grip_layout.setContentsMargins(0, 0, 0, 0)
        grip_layout.addWidget(grip)

        layout.addWidget(splitter, 1)
        layout.addWidget(self._build_right_panel())

    def _build_table_panel(self):
        panel = QFrame()
        panel.setStyleSheet("background: white; border-radius: 4px;")
        outer = QVBoxLayout(panel)
        outer.setContentsMargins(8, 8, 8, 4)
        outer.setSpacing(4)

        tools = QHBoxLayout()
        self.date_picker = AntdDateRangePicker(self.start_date, self.end_date)
        self.date_picker.valueChanged.connect(self._on_antd_date_changed)
        tools.addWidget(self.date_picker)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("查询操作员/产品码")
        self.name_input.setFixedWidth(160)
        self.name_input.setStyleSheet(
            "border: 1px solid #ddd; border-radius: 4px; padding: 4px 8px; background: white;"
        )
        tools.addWidget(self.name_input)
        tools.addStretch()

        self.code_lbl = QLabel("产品码：XXXX-XXXX-XXXX-XXXX")
        self.code_lbl.setStyleSheet(
            "font-size: 18px; font-weight: 900; font-family: 'YouSheBiaoTiHei';"
        )
        tools.addWidget(self.code_lbl)
        outer.addLayout(tools)

        self.table = QTableWidget(0, 22)
        groups = [
            {"label": "ID", "span": 1},
            {"label": "时间", "span": 1},
            {"label": "产品码", "span": 1},
            {"label": "操作员", "span": 1},
            {"label": "气密结果", "span": 1},
            {
                "label": "销钉1",
                "span": 4,
                "subs": ["压力值", "压力结果", "距离值", "距离结果"],
            },
            {
                "label": "销钉2",
                "span": 4,
                "subs": ["压力值", "压力结果", "距离值", "距离结果"],
            },
            {
                "label": "销钉3",
                "span": 4,
                "subs": ["压力值", "压力结果", "距离值", "距离结果"],
            },
            {
                "label": "销钉4",
                "span": 4,
                "subs": ["压力值", "压力结果", "距离值", "距离结果"],
            },
        ]
        groups.append({"label": "曲线图", "span": 1})
        self.group_header = GroupHeaderWidget(self.table, groups)
        outer.addWidget(self.group_header)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)
        header.setMinimumSectionSize(70)
        header.hide()

        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(36)
        self.table.setShowGrid(False)
        self.table.setMinimumHeight(120)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setStyleSheet(
            """
            QTableWidget { background: white; border: none; }
            QTableWidget::item { border-bottom: 1px solid #F0F0F0; padding: 4px; }
            QTableWidget::item:selected { background: #E6F4FF; color: #333; }
            """
        )

        column_widths = [
            56, 146, 190, 110, 86,
            84, 84, 84, 84,
            84, 84, 84, 84,
            84, 84, 84, 84,
            84, 84, 84, 84,
            92,
        ]
        for index, width in enumerate(column_widths):
            self.table.setColumnWidth(index, width)

        self._fill_table([])
        outer.addWidget(self.table, 1)

        footer = QHBoxLayout()
        self._export_btn = QPushButton("导出")
        self._export_btn.setStyleSheet(
            f"background: {self.PRIMARY}; color: white; padding: 6px 20px; border-radius: 4px; font-weight: bold; border: none;"
        )
        self._export_btn.clicked.connect(self._on_export_clicked)
        footer.addWidget(self._export_btn)
        footer.addStretch()

        self.prev_page_btn = self._create_page_button("<")
        self.prev_page_btn.clicked.connect(self._go_prev_page)
        footer.addWidget(self.prev_page_btn)

        self.page_buttons = []
        for _ in range(5):
            button = self._create_page_button("")
            button.clicked.connect(self._on_page_button_clicked)
            self.page_buttons.append(button)
            footer.addWidget(button)

        self.next_page_btn = self._create_page_button(">")
        self.next_page_btn.clicked.connect(self._go_next_page)
        footer.addWidget(self.next_page_btn)
        outer.addLayout(footer)

        return panel

    def _build_chart_panel(self):
        frame = QFrame()
        frame.setStyleSheet("background: #0f1429; border-radius: 4px;")
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)

        self.plot = pg.PlotWidget()
        self.plot.setBackground("#0f1429")
        self.plot.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._setup_chart()
        layout.addWidget(self.plot)
        return frame

    def _setup_chart(self):
        plot_item = self.plot.getPlotItem()
        plot_item.hideAxis("top")
        plot_item.hideAxis("right")

        left = plot_item.getAxis("left")
        left.setPen(QColor("#5d7bb9"))
        left.setTextPen(QColor("#9dbde3"))
        left.setLabel("力矩 (Nm)", color="#9dbde3")

        bottom = plot_item.getAxis("bottom")
        bottom.setPen(QColor("#5d7bb9"))
        bottom.setTextPen(QColor("#9dbde3"))
        bottom.setLabel("相对时间 (s)", color="#9dbde3")

        self.plot.showGrid(x=True, y=True, alpha=0.15)
        self.plot.setYRange(0, 100)
        self.plot.setXRange(0, 10)

        self.warn_line = pg.InfiniteLine(
            pos=self.threshold,
            angle=0,
            pen=pg.mkPen("#ff7676", width=2, style=Qt.PenStyle.DashLine),
        )
        self.plot.addItem(self.warn_line)

        menu = plot_item.getViewBox().menu
        menu.addSeparator()
        action = QAction("Set Safety Threshold...", self)
        action.triggered.connect(self._on_set_threshold)
        menu.addAction(action)
        self.curve_1 = self.plot.plot([], [], pen=pg.mkPen("#0958d9", width=2.5))
        self.curve_1_base = self.plot.plot([], [], pen=None)
        self.curve_1_fill = pg.FillBetweenItem(
            self.curve_1,
            self.curve_1_base,
            brush=QColor(9, 88, 217, 40),
        )
        self.plot.addItem(self.curve_1_fill)

        self.curve_2 = self.plot.plot([], [], pen=pg.mkPen("#389e0d", width=2.5))
        self.curve_2_base = self.plot.plot([], [], pen=None)
        self.curve_2_fill = pg.FillBetweenItem(
            self.curve_2,
            self.curve_2_base,
            brush=QColor(56, 158, 13, 30),
        )
        self.plot.addItem(self.curve_2_fill)

    def _on_set_threshold(self):
        keypad = NumericKeypad(str(self.threshold), self.window())
        if keypad.exec():
            try:
                self.threshold = float(keypad.get_value())
                self.warn_line.setPos(self.threshold)
            except ValueError:
                pass

    def _sample_torque_from_manual_store(self):
        """manual_store 数据变化时采样，仅在 COLLECTING 状态下采集。"""
        if self.manual_store is None:
            return
        if self._chart_state != "COLLECTING":
            return

        torque_1 = float(self.manual_store.torque_data.get("torque_1", 0.0))
        torque_2 = float(self.manual_store.torque_data.get("torque_2", 0.0))
        self._append_torque_sample(torque_1, torque_2, monotonic())

    def _append_torque_sample(self, torque_1: float, torque_2: float, sample_at: float):
        """COLLECTING 状态下累积采样点，超过上限时等间距降采样。"""
        if self._chart_started_at is None:
            self._chart_started_at = sample_at

        relative_time = max(0.0, sample_at - self._chart_started_at)
        self._chart_x.append(relative_time)
        self._chart_torque_1.append(float(torque_1))
        self._chart_torque_2.append(float(torque_2))

        # 超过上限时等间距降采样（保留首尾 + 均匀抽取）
        if len(self._chart_x) > self._chart_max_samples:
            self._downsample()

        self._refresh_torque_chart()

    def _downsample(self):
        """等间距降采样到 max_samples 的 75%，保留首尾点。"""
        target = int(self._chart_max_samples * 0.75)
        n = len(self._chart_x)
        if n <= target:
            return
        # 保留首尾 + 均匀抽取
        indices = [0]
        step = (n - 1) / (target - 1)
        for i in range(1, target - 1):
            indices.append(int(round(i * step)))
        indices.append(n - 1)
        self._chart_x = [self._chart_x[i] for i in indices]
        self._chart_torque_1 = [self._chart_torque_1[i] for i in indices]
        self._chart_torque_2 = [self._chart_torque_2[i] for i in indices]

    def _refresh_torque_chart(self):
        x_data = list(self._chart_x)
        y1_data = list(self._chart_torque_1)
        y2_data = list(self._chart_torque_2)
        baseline = [0.0] * len(x_data)

        self.curve_1.setData(x_data, y1_data)
        self.curve_1_base.setData(x_data, baseline)
        self.curve_2.setData(x_data, y2_data)
        self.curve_2_base.setData(x_data, baseline)

        # X 轴自适应: 始终从 0 开始，结尾自动扩展
        max_x = x_data[-1] if x_data else 10.0
        self.plot.setXRange(0, max(10.0, max_x * 1.05), padding=0)

        max_value = max(
            [self.threshold, 10.0] + y1_data + y2_data
        )
        self.plot.setYRange(0, max(100.0, max_value * 1.1), padding=0)

    # ------ 工序状态机 ------

    def _on_run_seq_changed(self, run_seq: int):
        """监听 VW22 (run_seq) 变化，驱动采集状态切换。"""
        if run_seq > 10 and self._chart_state != "COLLECTING":
            self._start_collecting()

    def _start_collecting(self):
        """进入 COLLECTING 状态：清空旧数据，准备新工序采集。"""
        self._chart_state = "COLLECTING"
        self._chart_started_at = None
        self._chart_x.clear()
        self._chart_torque_1.clear()
        self._chart_torque_2.clear()
        self._refresh_torque_chart()
        print(f"[GasPage] 工序采集开始")

    def _on_clear_statistics_clicked(self):
        if self.store is not None:
            self.store.reset_statistics()

    def _build_right_panel(self):
        panel = QFrame()
        panel.setFixedWidth(180)
        panel.setStyleSheet("background: transparent;")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.status_frame = QFrame()
        self.status_frame.setFixedHeight(80)
        self.status_layout = QVBoxLayout(self.status_frame)
        self.status_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_icon_label = QLabel("--")
        self.status_icon_label.setStyleSheet("color: white; font-size: 20px; font-weight: bold;")
        self.status_icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_text_label = QLabel("待检测")
        self.status_text_label.setStyleSheet(
            "color: white; font-size: 22px; font-weight: 900; font-family: 'YouSheBiaoTiHei';"
        )
        self.status_text_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_layout.addWidget(self.status_icon_label)
        self.status_layout.addWidget(self.status_text_label)
        layout.addWidget(self.status_frame)

        bars_frame = QFrame()
        bars_frame.setStyleSheet("background: white; border-radius: 6px;")
        bars_layout = QHBoxLayout(bars_frame)
        bars_layout.setContentsMargins(8, 8, 8, 8)
        bars_layout.setSpacing(8)

        ok_column = QVBoxLayout()
        ok_column.setSpacing(4)
        self.ok_bar = VerticalProgressBar(0, self.OK_GREEN, self.GREY_BG)
        ok_column.addWidget(self.ok_bar, 1)
        ok_title = QLabel("合格")
        ok_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ok_count_label = QLabel("0")
        self.ok_count_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ok_count_label.setStyleSheet(
            f"color: {self.OK_GREEN}; font-size: 18px; font-weight: bold; font-family: 'YouSheBiaoTiHei';"
        )
        ok_column.addWidget(ok_title)
        ok_column.addWidget(self.ok_count_label)
        bars_layout.addLayout(ok_column)

        ng_column = QVBoxLayout()
        ng_column.setSpacing(4)
        self.ng_bar = VerticalProgressBar(0, self.PROGRESS_NG, self.GREY_BG)
        ng_column.addWidget(self.ng_bar, 1)
        ng_title = QLabel("不合格")
        ng_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ng_count_label = QLabel("0")
        self.ng_count_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ng_count_label.setStyleSheet(
            f"color: {self.PROGRESS_NG}; font-size: 18px; font-weight: bold; font-family: 'YouSheBiaoTiHei';"
        )
        ng_column.addWidget(ng_title)
        ng_column.addWidget(self.ng_count_label)
        bars_layout.addLayout(ng_column)

        layout.addWidget(bars_frame, 1)

        clear_btn = QPushButton("统计清零")
        clear_btn.setFixedHeight(48)
        clear_btn.setStyleSheet(
            f"background: {self.PRIMARY}; color: white; border: none; border-radius: 6px; font-size: 18px; font-weight: bold; font-family: 'YouSheBiaoTiHei';"
        )
        layout.addWidget(clear_btn)
        self._set_status_card(0)
        return panel

    def _connect_signals(self):
        if not self.store:
            return

        self.name_input.editingFinished.connect(self._sync_search_to_store)
        self.store.dataChanged.connect(self._update_ui_from_store)
        self.store.recordInserted.connect(self._on_record_inserted)
        self._clear_statistics_button = self.findChildren(QPushButton)[-1]
        self._clear_statistics_button.clicked.connect(self._on_clear_statistics_clicked)
        if self.manual_store is not None and hasattr(self.manual_store, "dataChanged"):
            self.manual_store.dataChanged.connect(self._sample_torque_from_manual_store)
        self.store.runSeqChanged.connect(self._on_run_seq_changed)

    def _sync_search_to_store(self):
        if not self.store:
            return

        self.store.set_search(
            keyword=self.name_input.text().strip(),
            date_start=self.start_date.toString("yyyy-MM-dd"),
            date_end=self.end_date.toString("yyyy-MM-dd"),
        )

    def _on_antd_date_changed(self, start, end):
        self.start_date = start
        self.end_date = end
        if self.store:
            self._sync_search_to_store()

    def _update_ui_from_store(self):
        if not self.store:
            return

        if self.name_input.text() != self.store.search_keyword:
            self.name_input.setText(self.store.search_keyword)

        start_text, end_text = self.store.date_range
        if start_text and end_text:
            start = QDate.fromString(start_text, "yyyy-MM-dd")
            end = QDate.fromString(end_text, "yyyy-MM-dd")
            if start.isValid() and end.isValid():
                self.start_date = start
                self.end_date = end
                self.date_picker.start_date = start
                self.date_picker.end_date = end
                self.date_picker.lbl_start.setText(start.toString("yyyy-MM-dd"))
                self.date_picker.lbl_end.setText(end.toString("yyyy-MM-dd"))

        product_code = self.store.current_data.get("product_code") or "XXXX-XXXX-XXXX-XXXX"
        self.code_lbl.setText(f"产品码：{product_code}")
        self._set_status_card(self.store.get_status_card_state())
        self._reload_records_if_needed()

        ok_count = int(self.store.current_data.get("count_ok", 0))
        ng_count = int(self.store.current_data.get("count_ng", 0))
        self._update_statistics(ok_count, ng_count)

    def _on_record_inserted(self, record_id: int, product_code: str, timestamp: str):
        # 进入 FROZEN 状态：冻结当前曲线用于截图
        self._chart_state = "FROZEN"
        print(f"[GasPage] 工序采集冻结 (record_id={record_id})")

        relative_path = ""
        absolute_path = None
        try:
            relative_path, absolute_path = self._build_chart_image_path(
                record_id, product_code, timestamp
            )
            absolute_path.parent.mkdir(parents=True, exist_ok=True)
            if not self.plot.grab().save(str(absolute_path), "PNG"):
                self._log_chart_error(
                    "save_chart",
                    record_id=record_id,
                    image_path=relative_path,
                    error="QPixmap.save returned False",
                )
            else:
                try:
                    self.store.update_chart_image(record_id, relative_path)
                except Exception as exc:
                    self._log_chart_error(
                        "update_chart_image",
                        record_id=record_id,
                        image_path=relative_path,
                        error=exc,
                    )
        except Exception as exc:
            self._log_chart_error(
                "save_chart",
                record_id=record_id,
                image_path=relative_path or absolute_path,
                error=exc,
            )
        self._reload_records_if_needed(force=True)

    def _reload_records_if_needed(self, force: bool = False):
        if not self.store:
            return

        query_state = (
            self.store.search_keyword,
            self.store.date_range,
            self.store.page_index,
            self.store.page_size,
        )

        if force or query_state != self._last_query_state:
            records, total_count = self.store.load_records()
            self._fill_table(records)
            self._refresh_pagination(total_count)
            self._last_query_state = query_state

    def _fill_table(self, records: list[dict]):
        for row in range(self.table.rowCount()):
            widget = self.table.cellWidget(row, self.IMAGE_COLUMN)
            if widget is not None:
                widget.deleteLater()
        self.table.clearContents()
        self.table.setRowCount(len(records))

        result_cols = {4, 6, 8, 10, 12, 14, 16, 18, 20}
        gas_result_col = 4

        for row, record in enumerate(records):
            row_data = [
                str(record.get("id", "")),
                self._format_timestamp(record.get("timestamp", "")),
                record.get("product_code", ""),
                record.get("operator", ""),
                GasStore.result_to_text(record.get("airtight", 0)),
                self._format_float(record.get("pin1_pressure")),
                GasStore.result_to_text(record.get("pin1_pressure_result", 0)),
                self._format_float(record.get("pin1_distance")),
                GasStore.result_to_text(record.get("pin1_distance_result", 0)),
                self._format_float(record.get("pin2_pressure")),
                GasStore.result_to_text(record.get("pin2_pressure_result", 0)),
                self._format_float(record.get("pin2_distance")),
                GasStore.result_to_text(record.get("pin2_distance_result", 0)),
                self._format_float(record.get("pin3_pressure")),
                GasStore.result_to_text(record.get("pin3_pressure_result", 0)),
                self._format_float(record.get("pin3_distance")),
                GasStore.result_to_text(record.get("pin3_distance_result", 0)),
                self._format_float(record.get("pin4_pressure")),
                GasStore.result_to_text(record.get("pin4_pressure_result", 0)),
                self._format_float(record.get("pin4_distance")),
                GasStore.result_to_text(record.get("pin4_distance_result", 0)),
            ]

            is_gas_ng = row_data[gas_result_col] == "NG"
            for col, value in enumerate(row_data):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                if is_gas_ng:
                    item.setForeground(QColor(self.NG_RED))
                elif col in result_cols and value == "NG":
                    item.setForeground(QColor(self.NG_RED))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                elif col in result_cols and value == "OK":
                    item.setForeground(QColor(self.OK_GREEN))
                elif col in result_cols and value == "未记录":
                    item.setForeground(QColor("#BFBFBF"))

                self.table.setItem(row, col, item)

            image_path = record.get("chart_image") or ""
            image_exists = bool(image_path) and self._resolve_image_path(image_path).is_file()
            self.table.setCellWidget(
                row,
                self.IMAGE_COLUMN,
                self._create_chart_button(image_path, enabled=image_exists),
            )

    def _create_chart_button(self, image_path: str, enabled: bool) -> QPushButton:
        button = QPushButton("查看" if enabled else "无图")
        button.setEnabled(enabled)
        if enabled:
            button.setStyleSheet(
                f"QPushButton {{ background: {self.PRIMARY}; color: white; border: none; "
                f"border-radius: 4px; padding: 4px 12px; margin: 4px 10px; }}"
            )
            button.clicked.connect(lambda checked=False, path=image_path: self._show_chart_image(path))
        else:
            button.setStyleSheet(
                "QPushButton { background: #F5F5F5; color: #BFBFBF; border: 1px solid #E8E8E8; "
                "border-radius: 4px; padding: 4px 12px; margin: 4px 10px; }"
            )
        return button

    def _show_chart_image(self, path: str) -> None:
        absolute_path = self._resolve_image_path(path)
        if not absolute_path.is_file():
            self._log_chart_error("show_chart_image", image_path=path, error="image file not found")
            QMessageBox.information(self, "提示", "曲线图文件不存在或无法加载。")
            return

        pixmap = QPixmap(str(absolute_path))
        if pixmap.isNull():
            self._log_chart_error("show_chart_image", image_path=path, error="QPixmap is null")
            QMessageBox.information(self, "提示", "曲线图文件不存在或无法加载。")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("曲线图查看")
        dialog.setModal(True)
        dialog.resize(980, 720)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)

        image_label = QLabel(dialog)
        image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_label.setPixmap(
            pixmap.scaled(
                940,
                680,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
        )
        layout.addWidget(image_label)
        dialog.exec()

    def _build_chart_image_path(
        self, record_id: int, product_code: str, timestamp: str
    ) -> tuple[str, Path]:
        safe_product_code = self._sanitize_product_code(product_code)
        parsed_timestamp = self._parse_record_timestamp(timestamp)
        relative_path = (
            Path("data")
            / "image"
            / f"{int(record_id)}_{safe_product_code}_{parsed_timestamp.strftime('%Y-%m-%d_%H-%M')}.png"
        )
        return relative_path.as_posix(), self._project_root / relative_path

    @staticmethod
    def _sanitize_product_code(product_code: str) -> str:
        safe_value = re.sub(r"[^A-Za-z0-9_-]+", "_", (product_code or "").strip()).strip("_")
        return safe_value or "unknown"

    @staticmethod
    def _parse_record_timestamp(timestamp: str) -> datetime:
        try:
            return datetime.fromisoformat(timestamp)
        except ValueError:
            return datetime.now()

    def _resolve_image_path(self, stored_path: str | Path) -> Path:
        image_path = Path(stored_path)
        if image_path.is_absolute():
            return image_path
        return self._project_root / image_path

    def _log_chart_error(self, action: str, **context) -> None:
        log_dir = self._project_root / "log"
        log_dir.mkdir(parents=True, exist_ok=True)

        parts = [f"time={datetime.now().isoformat(timespec='seconds')}", f"action={action}"]
        for key, value in context.items():
            if value in (None, ""):
                continue
            if isinstance(value, Path):
                value = value.as_posix()
            elif isinstance(value, Exception):
                value = f"{type(value).__name__}: {value}"
            parts.append(f"{key}={str(value).replace(chr(10), ' | ')}")

        with (log_dir / "gas_page.log").open("a", encoding="utf-8") as handle:
            handle.write(" ".join(parts) + "\n")

    def _update_statistics(self, ok_count: int, ng_count: int):
        total_count = ok_count + ng_count
        if total_count > 0:
            ok_percent = round(ok_count / total_count * 100)
            ng_percent = 100 - ok_percent
        else:
            ok_percent = 0
            ng_percent = 0

        self.ok_bar.value = ok_percent
        self.ng_bar.value = ng_percent
        self.ok_bar.update()
        self.ng_bar.update()
        self.ok_count_label.setText(str(ok_count))
        self.ng_count_label.setText(str(ng_count))

    def _set_status_card(self, status_state: str):
        if status_state == "ok":
            background = self.OK_GREEN
            title = "合格"
            icon = "OK"
        elif status_state == "ng":
            background = self.NG_RED
            title = "不合格"
            icon = "NG"
        else:
            background = "#8C8C8C"
            title = "待检测"
            icon = "--"

        self.status_frame.setStyleSheet(f"background-color: {background}; border-radius: 6px;")
        self.status_icon_label.setText(icon)
        self.status_text_label.setText(title)

    def _create_page_button(self, text: str) -> QPushButton:
        button = QPushButton(text)
        button.setFixedSize(30, 30)
        button.setStyleSheet(
            "background: white; color: #666; border: 1px solid #ddd; border-radius: 2px;"
        )
        return button

    def _style_page_button(self, button: QPushButton, active: bool = False):
        if active:
            button.setStyleSheet(
                f"background: {self.PRIMARY}; color: white; border: none; border-radius: 2px; font-weight: bold;"
            )
        else:
            button.setStyleSheet(
                "background: white; color: #666; border: 1px solid #ddd; border-radius: 2px;"
            )

    def _refresh_pagination(self, total_count: int):
        if not self.store:
            return

        total_pages = max(1, math.ceil(total_count / self.store.page_size))
        current_page = min(self.store.page_index, total_pages)
        if current_page != self.store.page_index:
            self.store.set_page(current_page)
            return

        self.prev_page_btn.setEnabled(current_page > 1)
        self.next_page_btn.setEnabled(current_page < total_pages)

        start_page = max(1, current_page - 2)
        end_page = min(total_pages, start_page + len(self.page_buttons) - 1)
        start_page = max(1, end_page - len(self.page_buttons) + 1)

        pages = list(range(start_page, end_page + 1))
        while len(pages) < len(self.page_buttons):
            pages.append(None)

        for button, page_number in zip(self.page_buttons, pages):
            if page_number is None:
                button.setText("")
                button.setEnabled(False)
                self._style_page_button(button, active=False)
                continue

            button.setText(str(page_number))
            button.setEnabled(True)
            self._style_page_button(button, active=page_number == current_page)

    def _on_page_button_clicked(self):
        if not self.store:
            return

        button = self.sender()
        if not isinstance(button, QPushButton):
            return

        text = button.text().strip()
        if text.isdigit():
            self.store.set_page(int(text))

    def _go_prev_page(self):
        if self.store:
            self.store.set_page(max(1, self.store.page_index - 1))

    def _go_next_page(self):
        if self.store:
            self.store.set_page(self.store.page_index + 1)

    @staticmethod
    def _format_timestamp(value: str) -> str:
        if not value:
            return ""
        try:
            return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return value

    @staticmethod
    def _format_float(value) -> str:
        if value in (None, ""):
            return ""
        try:
            return f"{float(value):.3f}".rstrip("0").rstrip(".")
        except (TypeError, ValueError):
            return ""

    def _on_export_clicked(self):
        if not self.store:
            return

        records = self.store.load_all_records()
        if not records:
            QMessageBox.information(self, "提示", "当前时间范围内没有记录可导出。")
            return

        start_str = self.start_date.toString("yyyy-MM-dd")
        end_str = self.end_date.toString("yyyy-MM-dd")
        default_name = f"气检记录_{start_str}_{end_str}.xlsx"

        filepath, _ = QFileDialog.getSaveFileName(
            self, "导出气检记录", default_name, "Excel 文件 (*.xlsx)"
        )
        if not filepath:
            return

        try:
            self._export_to_excel(records, filepath)
            QMessageBox.information(self, "提示", f"导出成功！\n共 {len(records)} 条记录。")
        except Exception as exc:
            QMessageBox.warning(self, "导出失败", f"导出时发生错误：\n{exc}")

    @staticmethod
    def _export_to_excel(records: list[dict], filepath: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        headers = [
            "ID", "时间", "产品码", "操作员", "气密结果",
            "销钉1压力值", "销钉1压力结果", "销钉1距离值", "销钉1距离结果",
            "销钉2压力值", "销钉2压力结果", "销钉2距离值", "销钉2距离结果",
            "销钉3压力值", "销钉3压力结果", "销钉3距离值", "销钉3距离结果",
            "销钉4压力值", "销钉4压力结果", "销钉4距离值", "销钉4距离结果",
        ]
        fields = [
            "id", "timestamp", "product_code", "operator", "airtight",
            "pin1_pressure", "pin1_pressure_result", "pin1_distance", "pin1_distance_result",
            "pin2_pressure", "pin2_pressure_result", "pin2_distance", "pin2_distance_result",
            "pin3_pressure", "pin3_pressure_result", "pin3_distance", "pin3_distance_result",
            "pin4_pressure", "pin4_pressure_result", "pin4_distance", "pin4_distance_result",
        ]
        result_fields = {"airtight", "pin1_pressure_result", "pin1_distance_result",
                         "pin2_pressure_result", "pin2_distance_result",
                         "pin3_pressure_result", "pin3_distance_result",
                         "pin4_pressure_result", "pin4_distance_result"}

        wb = Workbook()
        ws = wb.active
        ws.title = "气检记录"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="597EF7", end_color="597EF7", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(fields, 1):
                value = record.get(field, "")
                if field in result_fields:
                    value = GasStore.result_to_text(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center

        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
                max_length + 4, 40
            )

        wb.save(filepath)
