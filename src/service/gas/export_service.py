from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.store.gas_store import GasStore


def export_gas_records_to_excel(records: list[dict], filepath: str) -> None:
    headers = [
        "ID", "时间", "产品码", "操作员", "气密结果",
        "销钉1压力值", "销钉1压力结果", "销钉1距离值", "销钉1距离结果",
        "销钉2压力值", "销钉2压力结果", "销钉2距离值", "销钉2距离结果",
        "销钉3压力值", "销钉3压力结果", "销钉3距离值", "销钉3距离结果",
        "销钉4压力值", "销钉4压力结果", "销钉4距离值", "销钉4距离结果",
    ]
    fields = [
        "id", "timestamp", "product_code", "operator", "airtight",
        "pin1_pressure", "pin1_pressure_result", "pin1_distance", "pin1_distance_result",
        "pin2_pressure", "pin2_pressure_result", "pin2_distance", "pin2_distance_result",
        "pin3_pressure", "pin3_pressure_result", "pin3_distance", "pin3_distance_result",
        "pin4_pressure", "pin4_pressure_result", "pin4_distance", "pin4_distance_result",
    ]
    result_fields = {
        "airtight",
        "pin1_pressure_result",
        "pin1_distance_result",
        "pin2_pressure_result",
        "pin2_distance_result",
        "pin3_pressure_result",
        "pin3_distance_result",
        "pin4_pressure_result",
        "pin4_distance_result",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "气检记录"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="597EF7", end_color="597EF7", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(fields, 1):
            value = record.get(field, "")
            if field in result_fields:
                value = GasStore.result_to_text(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, ws.max_row + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_length + 4,
            40,
        )

    wb.save(filepath)
